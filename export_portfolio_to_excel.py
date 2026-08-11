"""
Export Portfolio Data to Excel Format
Generates synthetic portfolio data and exports to readable Excel file
"""

import sys
import os
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from data_connectors.mock_portfolio import MockPortfolioGenerator
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def export_portfolio_to_excel(filename='Portfolio_Data.xlsx'):
    """
    Generate synthetic portfolio data and export to Excel with formatting
    """
    print("🔄 Generating synthetic portfolio data...")
    portfolio = MockPortfolioGenerator.generate_portfolio()
    
    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ===== SHEET 1: PORTFOLIO SUMMARY =====
    print("📊 Creating Portfolio Summary sheet...")
    ws_summary = wb.create_sheet("Portfolio Summary", 0)
    
    summary_data = {
        'Metric': [
            'Portfolio Value',
            'Total Capital Required',
            'Capital Utilization',
            'Average RORAC',
            'Total Loss Ratio',
            'Expected Profit',
            'Diversification Score',
            'Treaty Count',
            'Generated Date'
        ],
        'Value': [
            f"${portfolio['portfolio_value']:,.2f}",
            f"${portfolio['total_capital']:,.2f}",
            f"{portfolio['capital_utilization']:.2f}%",
            f"{portfolio['avg_rorac']:.2f}%",
            f"{portfolio['total_loss_ratio']:.2%}",
            f"${portfolio['expected_profit']:,.2f}",
            f"{portfolio['diversification_score']:.2f}",
            f"{portfolio['treaty_count']}",
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
    }
    
    df_summary = pd.DataFrame(summary_data)
    
    # Write summary data
    ws_summary.append(['PORTFOLIO SUMMARY METRICS'])
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="003d82", end_color="003d82", fill_type="solid")
    ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 25
    
    ws_summary.append([])  # Blank row
    
    # Add headers
    headers = ['Metric', 'Value']
    ws_summary.append(headers)
    header_fill = PatternFill(start_color="0084d6", end_color="0084d6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for row_num, row in enumerate(df_summary.values, 4):
        for col_num, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if col_num == 1:
                cell.font = Font(bold=True)
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 25
    
    # ===== SHEET 2: LOB BREAKDOWN =====
    print("📈 Creating LOB Breakdown sheet...")
    ws_lob = wb.create_sheet("LOB Breakdown", 1)
    
    lob_data = {
        'Line of Business': list(portfolio['lob_breakdown'].keys()),
        'Premium ($M)': [f"${v/1e6:,.2f}" for v in portfolio['lob_breakdown'].values()],
        'Premium (%)': [f"{(v/portfolio['portfolio_value'])*100:.1f}%" for v in portfolio['lob_breakdown'].values()]
    }
    
    df_lob = pd.DataFrame(lob_data)
    
    ws_lob.append(['LINE OF BUSINESS BREAKDOWN'])
    ws_lob.merge_cells('A1:C1')
    ws_lob['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_lob['A1'].fill = PatternFill(start_color="003d82", end_color="003d82", fill_type="solid")
    ws_lob['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_lob.row_dimensions[1].height = 25
    
    ws_lob.append([])
    
    headers = df_lob.columns.tolist()
    ws_lob.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_lob.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, row in enumerate(df_lob.values, 4):
        for col_num, value in enumerate(row, 1):
            cell = ws_lob.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    ws_lob.column_dimensions['A'].width = 25
    ws_lob.column_dimensions['B'].width = 18
    ws_lob.column_dimensions['C'].width = 15
    
    # ===== SHEET 3: GEOGRAPHIC BREAKDOWN =====
    print("🗺️ Creating Geographic Breakdown sheet...")
    ws_geo = wb.create_sheet("Geographic Breakdown", 2)
    
    geo_data = {
        'Geography': list(portfolio['geography_breakdown'].keys()),
        'Premium ($M)': [f"${v/1e6:,.2f}" for v in portfolio['geography_breakdown'].values()],
        'Premium (%)': [f"{(v/portfolio['portfolio_value'])*100:.1f}%" for v in portfolio['geography_breakdown'].values()]
    }
    
    df_geo = pd.DataFrame(geo_data)
    
    ws_geo.append(['GEOGRAPHIC BREAKDOWN'])
    ws_geo.merge_cells('A1:C1')
    ws_geo['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_geo['A1'].fill = PatternFill(start_color="003d82", end_color="003d82", fill_type="solid")
    ws_geo['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_geo.row_dimensions[1].height = 25
    
    ws_geo.append([])
    
    headers = df_geo.columns.tolist()
    ws_geo.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_geo.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, row in enumerate(df_geo.values, 4):
        for col_num, value in enumerate(row, 1):
            cell = ws_geo.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    ws_geo.column_dimensions['A'].width = 25
    ws_geo.column_dimensions['B'].width = 18
    ws_geo.column_dimensions['C'].width = 15
    
    # ===== SHEET 4: ALL TREATIES =====
    print("📋 Creating Treaties Details sheet...")
    ws_treaties = wb.create_sheet("Treaties Details", 3)
    
    treaties = portfolio['treaties']
    df_treaties = pd.DataFrame(treaties)
    
    # Reorder and format columns
    column_order = [
        'treaty_id', 'lob', 'geography', 'treaty_type', 'premium', 'incurred_loss',
        'loss_ratio', 'capital_requirement', 'expected_profit', 'rorac',
        'ceded_premium', 'renewal_date', 'performance_status', 'rating'
    ]
    
    df_treaties = df_treaties[column_order]
    
    # Rename columns for readability
    column_names = {
        'treaty_id': 'Treaty ID',
        'lob': 'Line of Business',
        'geography': 'Geography',
        'treaty_type': 'Treaty Type',
        'premium': 'Premium ($)',
        'incurred_loss': 'Incurred Loss ($)',
        'loss_ratio': 'Loss Ratio',
        'capital_requirement': 'Capital Req ($)',
        'expected_profit': 'Expected Profit ($)',
        'rorac': 'RORAC (%)',
        'ceded_premium': 'Ceded Premium ($)',
        'renewal_date': 'Renewal Date',
        'performance_status': 'Performance',
        'rating': 'Rating'
    }
    
    df_treaties = df_treaties.rename(columns=column_names)
    
    ws_treaties.append(['PORTFOLIO TREATIES - DETAILED ANALYSIS'])
    ws_treaties.merge_cells(f'A1:{chr(64 + len(df_treaties.columns))}1')
    ws_treaties['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_treaties['A1'].fill = PatternFill(start_color="003d82", end_color="003d82", fill_type="solid")
    ws_treaties['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_treaties.row_dimensions[1].height = 25
    
    ws_treaties.append([])
    
    headers = df_treaties.columns.tolist()
    ws_treaties.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_treaties.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws_treaties.row_dimensions[3].height = 30
    
    # Add treaty data with formatting
    for row_num, row in enumerate(df_treaties.values, 4):
        for col_num, value in enumerate(row, 1):
            cell = ws_treaties.cell(row=row_num, column=col_num)
            
            # Format currency columns
            if col_num in [5, 6, 8, 9, 11]:  # Premium, Loss, Capital, Profit, Ceded
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '$#,##0.00'
                else:
                    cell.value = value
            # Format percentage columns
            elif col_num in [7, 10]:  # Loss Ratio, RORAC
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '0.00%' if col_num == 7 else '0.00'
                else:
                    cell.value = value
            else:
                cell.value = value
            
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
    
    # Set column widths
    widths = [15, 20, 18, 16, 14, 16, 12, 16, 16, 12, 16, 14, 14, 10]
    for i, width in enumerate(widths, 1):
        ws_treaties.column_dimensions[chr(64 + i)].width = width
    
    # Freeze header row
    ws_treaties.freeze_panes = 'A4'
    
    # ===== SHEET 5: PERFORMANCE DISTRIBUTION =====
    print("📊 Creating Performance Distribution sheet...")
    ws_perf = wb.create_sheet("Performance Distribution", 4)
    
    perf_data = {
        'Performance Status': list(portfolio['performance_distribution'].keys()),
        'Count': list(portfolio['performance_distribution'].values()),
        'Percentage': [f"{(v/sum(portfolio['performance_distribution'].values()))*100:.1f}%" 
                       for v in portfolio['performance_distribution'].values()]
    }
    
    df_perf = pd.DataFrame(perf_data)
    
    ws_perf.append(['PERFORMANCE DISTRIBUTION'])
    ws_perf.merge_cells('A1:C1')
    ws_perf['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_perf['A1'].fill = PatternFill(start_color="003d82", end_color="003d82", fill_type="solid")
    ws_perf['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_perf.row_dimensions[1].height = 25
    
    ws_perf.append([])
    
    headers = df_perf.columns.tolist()
    ws_perf.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_perf.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, row in enumerate(df_perf.values, 4):
        for col_num, value in enumerate(row, 1):
            cell = ws_perf.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_perf.column_dimensions['A'].width = 25
    ws_perf.column_dimensions['B'].width = 12
    ws_perf.column_dimensions['C'].width = 15
    
    # Save workbook
    output_path = os.path.join(os.path.dirname(__file__), filename)
    wb.save(output_path)
    
    print(f"\n✅ Portfolio data exported successfully!")
    print(f"📁 File saved to: {output_path}")
    print(f"\n📊 Summary:")
    print(f"   • Portfolio Value: ${portfolio['portfolio_value']:,.2f}")
    print(f"   • Total Treaties: {portfolio['treaty_count']}")
    print(f"   • Capital Utilization: {portfolio['capital_utilization']:.2f}%")
    print(f"   • Average RORAC: {portfolio['avg_rorac']:.2f}%")
    print(f"   • Diversification Score: {portfolio['diversification_score']:.2f}")
    print(f"\n📄 Excel sheets created:")
    print(f"   1. Portfolio Summary - Key metrics overview")
    print(f"   2. LOB Breakdown - Premium by Line of Business")
    print(f"   3. Geographic Breakdown - Premium by Geography")
    print(f"   4. Treaties Details - All 50 treaties with full metrics")
    print(f"   5. Performance Distribution - Treaty performance status")
    
    return output_path

if __name__ == '__main__':
    export_portfolio_to_excel()
